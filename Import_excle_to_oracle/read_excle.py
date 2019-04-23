# 读写2007 excel
import openpyxl
import cx_Oracle as oracle
import os
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'

file_2007 = '/Users/zhoulinhan/个人余额.xlsx'

wb = openpyxl.load_workbook(file_2007)

sheet = wb['Sheet1']
nrows = sheet.max_row
ncols = sheet.max_column
list =[]
for i in range(2,nrows):
 #创建List
 data = []
 for j in range(1,ncols):
     data.append(sheet.cell(column=j,row=i).value)
 list.append(data)

print(list)
db = oracle.connect('apps/apps@193.168.42.68:1522/prod')

cursor = db.cursor ()

cursor.executemany('insert into cux.tb_user values(:id,:n,:p)',list);

db.commit()

cursor.close();

db.close();